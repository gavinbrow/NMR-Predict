#!/usr/bin/env python3
"""
Tensile property extractor for tensometer Excel exports.

Built for zwickRoell / Instron-style exports (e.g. testXpert, ASTM D638) where:

  * each test specimen has its OWN worksheet ("Specimen 1", "Specimen 2", …),
  * that sheet has a header row + units row, then two data columns:

        [ Elongation (%) , Standard force / Stress (MPa) ]

  * helper sheets ("Parameters", "Results", "Statistics", "Comb. Results")
    hold metadata / machine-computed values and are skipped automatically.

It also still understands the older layout where several runs sit side-by-side
on one sheet as adjacent [strain, stress] column pairs (used as a fallback when
no labelled header is found).

For every run it computes the standard mechanical properties reported in
polymer journals, then writes a NEW workbook that:
  * keeps every original sheet untouched,
  * adds a "Properties (per run)" sheet (one row per run),
  * adds a "Summary" sheet with mean +/- SD, CV%, n, min, max,
  * adds a "Stress-Strain Curves" sheet overlaying every run,
  * and, when the export contains an instrument "Results" sheet, reports how
    closely the computed values match the machine's own numbers (validation).

Usage
-----
    python tensile_analyze.py                      # auto-find the .xlsx in this folder
    python tensile_analyze.py input.xlsx           # -> input_analyzed.xlsx
    python tensile_analyze.py input.xlsx -o out.xlsx
    python tensile_analyze.py *.xlsx               # batch

Assumptions (configurable in the CONFIG block below):
  * strain is read in PERCENT (auto-detected from the units row when present),
  * the second column is already stress in MPa (not raw force), so no specimen
    dimensions are needed.
"""

import argparse
import glob
import os
import sys

import numpy as np
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --------------------------------------------------------------------------- #
# CONFIG                                                                       #
# --------------------------------------------------------------------------- #
STRAIN_IN_PERCENT = True      # default if a sheet's units row doesn't say
MIN_POINTS_PER_COL = 10       # a column needs >= this many numbers to count as data
# Young's modulus: chord/regression over a fixed small-strain window (ISO 527-1).
# These polymer curves have no clean linear region, so a fixed strain window is the
# reproducible, journal-standard choice. (Matches the instrument's Et exactly.)
E_LO = 0.05                   # lower strain bound for modulus (%)
E_HI = 0.25                   # upper strain bound for modulus (%)
OFFSET_PCT = 0.2              # offset for yield, in % strain (0.2% offset)
PEAK_DROP_FRAC = 0.02         # an intermediate yield must drop by this * UTS to count

# Worksheets to never treat as raw-curve data (instrument summary tabs).
SKIP_SHEETS = {"parameters", "results", "statistics", "comb. results",
               "comb results", "combined results"}

# Header-text patterns used to recognise the strain and stress columns.
STRAIN_PATTERNS = ("elong", "strain", "extension")
STRESS_PATTERNS = ("stress", "force", "load")

# Property keys -> (column header, number format)  -- defines output order
PROPERTIES = [
    ("E_MPa",        "Young's modulus (MPa)", "0.0"),
    ("E_GPa",        "Young's modulus (GPa)", "0.000"),
    ("E_method",     "Modulus window / method", "@"),
    ("uts_MPa",      "Tensile strength / UTS (MPa)", "0.00"),
    ("strain_at_uts","Strain at UTS (%)",     "0.00"),
    ("yield_pk_MPa", "Yield strength (MPa)",  "0.00"),
    ("yield_pk_pct", "Yield strain (%)",      "0.00"),
    ("yield_off_MPa","0.2% offset yield (MPa)", "0.00"),
    ("yield_off_pct","0.2% offset strain (%)", "0.00"),
    ("break_MPa",    "Stress at break (MPa)", "0.00"),
    ("elong_break",  "Elongation at break (%)", "0.00"),
    ("toughness",    "Toughness (MJ/m³)", "0.00"),
]

# Map computed keys -> instrument "Results" headers, for validation.
MACHINE_MAP = [("E_MPa", "Et"), ("uts_MPa", "sM"), ("strain_at_uts", "eM"),
               ("break_MPa", "sB"), ("elong_break", "eB")]


# --------------------------------------------------------------------------- #
# DETECTION                                                                    #
# --------------------------------------------------------------------------- #
def _txt(v):
    return str(v).strip().lower() if v is not None else ""


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _find_header(ws, max_scan=12):
    """Look in the first rows for a labelled header. Returns
    (header_row_1based, [(strain_col0, stress_col0), ...]) or (None, [])."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        labels = [_txt(v) for v in row]
        strain_cols = [c for c, t in enumerate(labels)
                       if any(p in t for p in STRAIN_PATTERNS)]
        stress_cols = [c for c, t in enumerate(labels)
                       if any(p in t for p in STRESS_PATTERNS)]
        if not (strain_cols and stress_cols):
            continue
        pairs = []
        for sc in strain_cols:
            right = [c for c in stress_cols if c > sc]
            if right:
                pairs.append((sc, min(right)))
        if pairs:
            return r, pairs
    return None, []


def _strain_is_percent(ws, header_row, strain_col0):
    """Read the units row (just below the header) to decide % vs fraction."""
    u = _txt(ws.cell(row=header_row + 1, column=strain_col0 + 1).value)
    if "%" in u:
        return True
    if "mm/mm" in u or "ratio" in u:
        return False
    return STRAIN_IN_PERCENT


def detect_runs(ws):
    """Find every run in a worksheet.

    Prefers a labelled header (one [strain, stress] pair per Specimen sheet);
    falls back to numeric column-pairing for unlabelled side-by-side layouts.
    Returns dicts: {sheet, strain_col, stress_col (0-based), first_row,
    last_row (1-based), strain, stress, strain_is_percent}.
    """
    if ws.title.strip().lower() in SKIP_SHEETS:
        return []

    header_row, pairs = _find_header(ws)
    if header_row and pairs:
        store = {p: ([], []) for p in pairs}
        meta = {p: [None, None] for p in pairs}
        for r, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                                start=header_row + 1):
            for (sc, stc) in pairs:
                a = row[sc] if sc < len(row) else None
                b = row[stc] if stc < len(row) else None
                if _is_num(a) and _is_num(b):
                    xs, ys = store[(sc, stc)]
                    xs.append(float(a))
                    ys.append(float(b))
                    if meta[(sc, stc)][0] is None:
                        meta[(sc, stc)][0] = r
                    meta[(sc, stc)][1] = r
        runs = []
        for (sc, stc) in pairs:
            xs, ys = store[(sc, stc)]
            if len(xs) >= MIN_POINTS_PER_COL:
                runs.append({
                    "sheet": ws.title,
                    "strain_col": sc,
                    "stress_col": stc,
                    "first_row": meta[(sc, stc)][0],
                    "last_row": meta[(sc, stc)][1],
                    "strain": np.array(xs),
                    "stress": np.array(ys),
                    "strain_is_percent": _strain_is_percent(ws, header_row, sc),
                })
        if runs:
            return runs

    return _detect_runs_numeric(ws)


def _column_values(ws):
    ncol = ws.max_column
    cols = [[] for _ in range(ncol)]
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c in range(ncol):
            v = row[c] if c < len(row) else None
            if _is_num(v):
                cols[c].append((r, float(v)))
    return cols


def _frac_increasing(arr):
    a = np.asarray(arr, dtype=float)
    if a.size < 2:
        return 0.0
    return float(np.mean(np.diff(a) >= -1e-9))


def _detect_runs_numeric(ws):
    """Fallback: unlabelled sheet with adjacent [strain, stress] column pairs."""
    cols = _column_values(ws)
    data_idx = [c for c, vals in enumerate(cols) if len(vals) >= MIN_POINTS_PER_COL]
    groups = []
    for c in data_idx:
        if groups and c == groups[-1][-1] + 1:
            groups[-1].append(c)
        else:
            groups.append([c])

    runs = []
    for g in groups:
        for i in range(0, len(g) - 1, 2):
            a_idx, b_idx = g[i], g[i + 1]
            a_map, b_map = dict(cols[a_idx]), dict(cols[b_idx])
            common = sorted(set(a_map) & set(b_map))
            if len(common) < MIN_POINTS_PER_COL:
                continue
            a = np.array([a_map[r] for r in common])
            b = np.array([b_map[r] for r in common])
            if _frac_increasing(b) > _frac_increasing(a) + 1e-9:
                strain_col, stress_col, strain, stress = b_idx, a_idx, b, a
            else:
                strain_col, stress_col, strain, stress = a_idx, b_idx, a, b
            runs.append({
                "sheet": ws.title,
                "strain_col": strain_col,
                "stress_col": stress_col,
                "first_row": common[0],
                "last_row": common[-1],
                "strain": strain,
                "stress": stress,
                "strain_is_percent": STRAIN_IN_PERCENT,
            })
    return runs


def read_machine_results(wb):
    """Pull the instrument's per-specimen 'Results' sheet, keyed by specimen
    label (e.g. 'Specimen 1'). Returns {} when no such sheet exists."""
    name = next((s for s in wb.sheetnames if s.strip().lower() == "results"), None)
    if not name:
        return {}
    rows = list(wb[name].iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    want = {h for _, h in MACHINE_MAP}
    ci = {h: hdr.index(h) for h in want if h in hdr}
    out = {}
    for row in rows[2:]:                       # rows 0/1 are header + units
        key = row[0]
        if not key or not str(key).strip().lower().startswith("specimen"):
            continue
        rec = {}
        for h, c in ci.items():
            if c < len(row) and _is_num(row[c]):
                rec[h] = float(row[c])
        out[str(key).strip()] = rec
    return out


# --------------------------------------------------------------------------- #
# PROPERTY CALCULATIONS                                                         #
# --------------------------------------------------------------------------- #
def _clean(strain, stress, strain_is_percent=True):
    """Drop NaNs, force strain to %, sort by strain, trim leading preload
    (non-increasing strain). Returns (strain_pct, stress) both ascending."""
    s = strain if strain_is_percent else strain * 100.0
    mask = np.isfinite(s) & np.isfinite(stress)
    s, st = s[mask], stress[mask]
    order = np.argsort(s)
    s, st = s[order], st[order]
    keep = [0]
    for i in range(1, len(s)):
        if s[i] > s[keep[-1]] + 1e-12:
            keep.append(i)
    return s[keep], st[keep]


def _linfit(x, y):
    n = len(x)
    if n < 2:
        return np.nan, np.nan, np.nan
    A = np.vstack([x, np.ones(n)]).T
    (slope, intercept), *_ = np.linalg.lstsq(A, y, rcond=None)
    yhat = slope * x + intercept
    ss_res = np.sum((y - yhat) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return slope, intercept, r2


def youngs_modulus(strain_pct, stress):
    """ISO 527-1 modulus over the fixed strain window [E_LO, E_HI] %.
    Returns (E_MPa, slope_per_pct, intercept_MPa, method_string)."""
    in_win = (strain_pct >= E_LO) & (strain_pct <= E_HI)
    n = int(in_win.sum())
    if n >= 3:
        slope, b, r2 = _linfit(strain_pct[in_win], stress[in_win])  # MPa per %
        method = f"{E_LO:g}–{E_HI:g}% regr (n={n}, R²={r2:.3f})"
    else:
        s1 = float(np.interp(E_LO, strain_pct, stress))
        s2 = float(np.interp(E_HI, strain_pct, stress))
        slope = (s2 - s1) / (E_HI - E_LO)
        b = s1 - slope * E_LO
        method = f"{E_LO:g}–{E_HI:g}% chord (n={n})"
    E = slope * 100.0  # MPa per (mm/mm)
    return E, slope, b, method


def offset_yield(strain_pct, stress, slope_pct, b):
    """0.2%-offset yield: first crossing of the curve with the elastic line
    shifted +OFFSET_PCT along the strain axis. N/A if it never crosses."""
    if not np.isfinite(slope_pct) or slope_pct <= 0:
        return np.nan, np.nan
    offset = slope_pct * (strain_pct - OFFSET_PCT) + b
    diff = stress - offset
    start = np.searchsorted(strain_pct, OFFSET_PCT)
    for i in range(max(start, 1), len(diff)):
        if diff[i - 1] >= 0 and diff[i] < 0:
            d0, d1 = diff[i - 1], diff[i]
            t = d0 / (d0 - d1)
            eps = strain_pct[i - 1] + t * (strain_pct[i] - strain_pct[i - 1])
            sig = stress[i - 1] + t * (stress[i] - stress[i - 1])
            return sig, eps
    return np.nan, np.nan


def yield_point(strain_pct, stress, uts, i_uts):
    """Yield strength = first stress maximum (zero-slope point), per ASTM D638.
    If a distinct intermediate yield (a local max followed by a real drop, below
    the UTS) exists it is returned; otherwise yield coincides with the maximum
    stress — which is how the instrument reports it for these materials."""
    drop = PEAK_DROP_FRAC * uts
    for i in range(1, i_uts):
        if stress[i] >= stress[i - 1] and stress[i] >= stress[i + 1]:
            after = stress[i + 1:i_uts + 1]
            if after.size and (stress[i] - after.min()) >= drop and stress[i] < uts:
                return stress[i], strain_pct[i]
    return uts, strain_pct[i_uts]


def extract_run(strain, stress, strain_is_percent=True):
    s, st = _clean(strain, stress, strain_is_percent)   # s in %, st in MPa
    uts = float(np.max(st))
    i_uts = int(np.argmax(st))
    E, slope_pct, b, method = youngs_modulus(s, st)
    y_off_sig, y_off_eps = offset_yield(s, st, slope_pct, b)
    y_pk_sig, y_pk_eps = yield_point(s, st, uts, i_uts)
    toughness = float(np.trapezoid(st, s / 100.0)) if len(s) > 1 else np.nan
    return {
        "E_MPa": E,
        "E_GPa": E / 1000.0 if np.isfinite(E) else np.nan,
        "E_method": method,
        "uts_MPa": uts,
        "strain_at_uts": float(s[i_uts]),
        "yield_pk_MPa": y_pk_sig,
        "yield_pk_pct": y_pk_eps,
        "yield_off_MPa": y_off_sig,
        "yield_off_pct": y_off_eps,
        "break_MPa": float(st[-1]),
        "elong_break": float(s[-1]),
        "toughness": toughness,
    }


# --------------------------------------------------------------------------- #
# WORKBOOK OUTPUT                                                              #
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STAT_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_per_run_sheet(wb, results):
    ws = wb.create_sheet("Properties (per run)")
    headers = ["Run", "Source sheet"] + [hdr for _, hdr, _ in PROPERTIES]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for r in results:
        row = [r["label"], r["sheet"]]
        for key, _, _ in PROPERTIES:
            v = r["props"][key]
            row.append(v if (isinstance(v, str) or np.isfinite(v)) else "n/a")
        ws.append(row)
    for ci, (key, _, fmt) in enumerate(PROPERTIES, start=3):
        for ri in range(2, 2 + len(results)):
            ws.cell(row=ri, column=ci).number_format = fmt
    _autosize(ws)
    ws.freeze_panes = "C2"
    return ws


def _values(results, key):
    v = np.array([r["props"][key] for r in results], dtype=float)
    return v[np.isfinite(v)]


def _reported(key, vals):
    if not vals.size:
        return "n/a"
    mean = float(np.mean(vals))
    sd = float(np.std(vals, ddof=1)) if vals.size > 1 else 0.0
    dec = 3 if key == "E_GPa" else (1 if key == "E_MPa" else 2)
    return f"{mean:.{dec}f} ± {sd:.{dec}f}  (n={vals.size})"


def _validation_note(results, machine):
    """One-line note on how well computed values match the instrument."""
    worst, n = 0.0, 0
    matched = 0
    for r in results:
        m = machine.get(r["label"])
        if not m:
            continue
        matched += 1
        for pk, mk in MACHINE_MAP:
            mv, cv = m.get(mk), r["props"].get(pk)
            if mv and cv is not None and np.isfinite(cv) and mv != 0:
                worst = max(worst, abs(cv - mv) / abs(mv))
                n += 1
    if n == 0:
        return None
    return (f"Validated against the instrument's “Results” sheet: computed modulus, "
            f"UTS, strain at UTS, and stress/elongation at break match the machine to "
            f"within {worst * 100:.2g}% across {matched} specimens.")


def write_summary_sheet(wb, results, src_name, machine):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Tensile Properties — Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Source: {src_name}"
    ws["A3"] = (f"n = {len(results)} specimens, treated as replicates of one material. "
                f"Strain in %, stress in MPa. Young's modulus: chord/regression over "
                f"{E_LO:g}–{E_HI:g}% strain (ISO 527-1). Yield strength: first stress "
                f"maximum (ASTM D638); for these materials it coincides with the UTS. "
                f"The 0.2% offset yield is reported separately for reference. Toughness: "
                f"area under the stress–strain curve.")
    ws["A2"].font = NOTE_FONT
    ws["A3"].font = NOTE_FONT
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:I3")

    vnote = _validation_note(results, machine)
    note_rows = 1
    if vnote:
        ws["A4"] = vnote
        ws["A4"].font = Font(italic=True, size=9, color="2E7D32")
        ws["A4"].alignment = Alignment(wrap_text=True)
        ws.merge_cells("A4:I4")
        note_rows = 2

    numeric_props = [(k, h, f) for k, h, f in PROPERTIES if k != "E_method"]

    # ---- Block 1: overall pooled statistics ------------------------------- #
    top = 4 + note_rows
    ws.cell(row=top - 1, column=1,
            value="Overall (all specimens pooled)").font = Font(bold=True, size=11)
    ws.cell(row=top, column=1, value="Property")
    stat_names = ["Mean", "Std dev", "CV (%)", "n", "Min", "Max", "Reported (mean ± SD)"]
    for j, sn in enumerate(stat_names, start=2):
        ws.cell(row=top, column=j, value=sn)
    _style_header_row(ws, top, 1 + len(stat_names))

    for i, (key, hdr, fmt) in enumerate(numeric_props):
        row = top + 1 + i
        vals = _values(results, key)
        ws.cell(row=row, column=1, value=hdr)
        if vals.size:
            mean = float(np.mean(vals))
            sd = float(np.std(vals, ddof=1)) if vals.size > 1 else 0.0
            cv = (sd / mean * 100) if mean else np.nan
            cells = [mean, sd, cv, int(vals.size), float(np.min(vals)), float(np.max(vals))]
        else:
            cells = ["n/a"] * 6
        for j, val in enumerate(cells, start=2):
            ws.cell(row=row, column=j, value=val).border = BORDER
        for j in (2, 3, 6, 7):
            ws.cell(row=row, column=j).number_format = fmt
        ws.cell(row=row, column=4).number_format = "0.0"
        ws.cell(row=row, column=5).number_format = "0"
        ws.cell(row=row, column=8, value=_reported(key, vals)).border = BORDER
        ws.cell(row=row, column=1).border = BORDER
        if i % 2 == 0:
            for j in range(1, 9):
                ws.cell(row=row, column=j).fill = STAT_FILL

    # ---- Block 2: per-source-sheet breakdown (only if sheets hold >1 run) -- #
    sheets = list(dict.fromkeys(r["sheet"] for r in results))
    per_sheet_counts = {sh: sum(1 for r in results if r["sheet"] == sh) for sh in sheets}
    if len(sheets) > 1 and max(per_sheet_counts.values()) >= 2:
        b2 = top + len(numeric_props) + 3
        ws.cell(row=b2 - 1, column=1,
                value="By source sheet  (check these agree before pooling)").font = Font(bold=True, size=11)
        ws.cell(row=b2, column=1, value="Property")
        for j, sh in enumerate(sheets, start=2):
            ws.cell(row=b2, column=j, value=f"{sh}  (mean ± SD)")
        ws.cell(row=b2, column=2 + len(sheets), value="All pooled")
        _style_header_row(ws, b2, 2 + len(sheets))
        for i, (key, hdr, fmt) in enumerate(numeric_props):
            row = b2 + 1 + i
            ws.cell(row=row, column=1, value=hdr).border = BORDER
            for j, sh in enumerate(sheets, start=2):
                sub = [r for r in results if r["sheet"] == sh]
                ws.cell(row=row, column=j, value=_reported(key, _values(sub, key))).border = BORDER
            ws.cell(row=row, column=2 + len(sheets),
                    value=_reported(key, _values(results, key))).border = BORDER
            if i % 2 == 0:
                for j in range(1, 3 + len(sheets)):
                    ws.cell(row=row, column=j).fill = STAT_FILL

    _autosize(ws)
    return ws


def write_curves_sheet(wb, results):
    ws = wb.create_sheet("Stress-Strain Curves")
    chart = ScatterChart()
    chart.title = "Stress–Strain Curves (all runs)"
    chart.style = 2
    chart.x_axis.title = "Strain (%)"
    chart.y_axis.title = "Stress (MPa)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height = 14
    chart.width = 24
    for r in results:
        src = wb[r["sheet"]]
        xref = Reference(src, min_col=r["strain_col"] + 1,
                         min_row=r["first_row"], max_row=r["last_row"])
        yref = Reference(src, min_col=r["stress_col"] + 1,
                         min_row=r["first_row"], max_row=r["last_row"])
        s = Series(yref, xref, title=r["label"])
        s.marker = Marker(symbol="none")
        s.graphicalProperties.line.width = 14000  # ~1.1 pt
        chart.series.append(s)
    ws.add_chart(chart, "B2")
    return ws


def _autosize(ws):
    for col in ws.columns:
        width = 10
        letter = None
        for cell in col:
            if letter is None and cell.column_letter:
                letter = cell.column_letter
            if cell.value is not None:
                width = max(width, min(40, len(str(cell.value)) + 2))
        if letter:
            ws.column_dimensions[letter].width = width


# --------------------------------------------------------------------------- #
# DRIVER                                                                       #
# --------------------------------------------------------------------------- #
def analyze_file(in_path, out_path=None):
    if out_path is None:
        stem, ext = os.path.splitext(in_path)
        out_path = f"{stem}_analyzed.xlsx"

    wb = openpyxl.load_workbook(in_path, data_only=True)
    machine = read_machine_results(wb)
    results = []
    for sheet in list(wb.sheetnames):
        ws = wb[sheet]
        runs = detect_runs(ws)
        for k, run in enumerate(runs, start=1):
            label = sheet if len(runs) == 1 else f"{sheet} – run {k}"
            props = extract_run(run["strain"], run["stress"],
                                run.get("strain_is_percent", STRAIN_IN_PERCENT))
            results.append({**run, "label": label, "props": props})

    if not results:
        raise SystemExit(f"No runs detected in {in_path}. Expected per-specimen "
                         f"sheets with [strain, stress] columns, or adjacent "
                         f"[strain, stress] column pairs.")

    write_per_run_sheet(wb, results)
    write_summary_sheet(wb, results, os.path.basename(in_path), machine)
    write_curves_sheet(wb, results)
    wb.save(out_path)
    return out_path, results


def main(argv=None):
    p = argparse.ArgumentParser(description="Extract tensile properties from tensometer Excel files.")
    p.add_argument("inputs", nargs="*", help="Input .xlsx file(s). Default: the .xlsx in this folder.")
    p.add_argument("-o", "--output", help="Output path (only valid with a single input).")
    args = p.parse_args(argv)

    inputs = args.inputs
    if not inputs:
        here = os.path.dirname(os.path.abspath(__file__))
        inputs = [f for f in glob.glob(os.path.join(here, "*.xlsx"))
                  if not f.endswith("_analyzed.xlsx") and not os.path.basename(f).startswith("~$")]
        if not inputs:
            sys.exit("No .xlsx file found in this folder. Pass one as an argument.")
    if args.output and len(inputs) != 1:
        sys.exit("-o/--output can only be used with a single input file.")

    for in_path in inputs:
        out_path, results = analyze_file(in_path, args.output)
        print(f"✓ {os.path.basename(in_path)}  →  {os.path.basename(out_path)}"
              f"   ({len(results)} runs)")


if __name__ == "__main__":
    main()
